import os
from PyQt6.QtWidgets import QApplication, QDialog, QLabel, QTextEdit, QPushButton, QVBoxLayout, QFileDialog, QHBoxLayout
from openpyxl import Workbook, load_workbook

def bgGui():
    dialog = QDialog()  # Create an instance of QDialog
    dialog.setWindowTitle("Music Configuration")
    dialog.setGeometry(400, 400, 400, 200)

    bg_music_name_label = QLabel("Bg Music Name:")
    bg_music_name_textedit = QTextEdit()

    select_path_label = QLabel("Select Path:")
    select_path_textedit = QTextEdit()

    open_file_button = QPushButton("Open File")
    open_file_button.clicked.connect(lambda: saveBgMusicDir(select_path_textedit))

    save_button = QPushButton("Save")
    save_button.clicked.connect(lambda: save_bg_music(bg_music_name_textedit, select_path_textedit, dialog))

    bg_music_name_textedit.setFixedSize(200, 30)
    select_path_textedit.setFixedSize(200, 30)
    open_file_button.setFixedWidth(90)
    save_button.setFixedWidth(90)

    layout = QVBoxLayout(dialog)
    name_layout = QHBoxLayout(dialog)
    name_layout.addWidget(bg_music_name_label)
    name_layout.addWidget(bg_music_name_textedit)
    layout.addLayout(name_layout)

    path_layout = QHBoxLayout(dialog)
    path_layout.addWidget(select_path_label)
    path_layout.addWidget(select_path_textedit)
    layout.addLayout(path_layout)

    file_button_layout = QHBoxLayout(dialog)
    file_button_layout.addStretch()
    file_button_layout.addWidget(open_file_button)
    layout.addLayout(file_button_layout)

    save_button_layout = QHBoxLayout(dialog)
    save_button_layout.addStretch()
    save_button_layout.addWidget(save_button)
    save_button_layout.addStretch()
    layout.addLayout(save_button_layout)

    dialog.exec()

def saveBgMusicDir(select_path_textedit):
    voicePath = ""

    file_dialog = QFileDialog()

    file_dialog.setNameFilter("MP3 files (*.mp3)")

    file_dialog.setOption(QFileDialog.Option.DontUseNativeDialog, True)
    file_dialog.setOption(QFileDialog.Option.ShowDirsOnly, True)

    file_paths, _ = file_dialog.getOpenFileNames(
        None,
        "Select MP3 Files",
        voicePath
    )

    save_file_path = ""
    if file_paths:
        save_file_path = file_paths[0].replace("\\", "//")
        select_path_textedit.setPlainText(save_file_path)
        # Reset the border style
        select_path_textedit.setStyleSheet("")

def save_bg_music(bg_music_name_textedit, select_path_textedit, dialog):
    bg_music_name = bg_music_name_textedit.toPlainText()
    select_path = select_path_textedit.toPlainText()

    if not select_path.lower().endswith('.mp3') or not os.path.exists(select_path):
        # File is not an MP3 or does not exist, show red border
        select_path_textedit.setStyleSheet("border: 1px solid red;")
    else:
        # File is an MP3 and exists, remove any previous styling
        select_path_textedit.setStyleSheet("")

        # Open the existing workbook
        try:
            wb = load_workbook("app/static/files/Background Music List.xlsx")
        except FileNotFoundError:
            # If the file doesn't exist, create a new workbook
            wb = Workbook()

        ws = wb.active

        # Find the last row with data
        last_row = ws.max_row + 1

        # Add the new data to the next row
        ws.cell(row=last_row, column=1, value=bg_music_name)
        ws.cell(row=last_row, column=2, value=select_path)

        # Save the workbook
        wb.save("app/static/files/Background Music List.xlsx")

        # Close the dialog
        dialog.accept()
        