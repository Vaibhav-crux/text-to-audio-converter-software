import openpyxl
import os
from PyQt6.QtCore import Qt

def on_combobox_changed(selected_item):
    wb = openpyxl.load_workbook('app/static/files/Background Music List.xlsx')
    ws = wb.active

    max_row = ws.max_row
    bgAudio = ""

    if selected_item == "None":
        bgAudio = ""
    else:
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=2, values_only=True):
            if row[0] == selected_item:
                corresponding_value = row[1]
                bgAudio = corresponding_value
                break

    files_directory = "app/static/files"
    api_key_file_path = os.path.join(files_directory, "Latest Background Music Used.txt")
    with open(api_key_file_path, 'w') as file:
        file.write(bgAudio)

def bgMusic(backgroundComboBox):
    wb = openpyxl.load_workbook('app/static/files/Background Music List.xlsx')
    ws = wb.active

    max_row = ws.max_row

    my_list = ["None"]
    my_list += [value[0] for value in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=1, values_only=True)]

    backgroundComboBox.addItems(my_list)
    backgroundComboBox.setEditable(True)
    backgroundComboBox.lineEdit().setAlignment(Qt.AlignmentFlag.AlignCenter)
    backgroundComboBox.currentIndexChanged.connect(lambda index: on_combobox_changed(backgroundComboBox.currentText()))

    return backgroundComboBox
